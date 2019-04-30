import mimetypes
import os
import re

from django.forms.models import model_to_dict
from django.shortcuts import HttpResponse
from docxtpl import DocxTemplate
from openpyxl import Workbook, load_workbook


def get_file(file_name, file_name_for_download):
    fp = open(file_name, "rb")
    response = HttpResponse(fp.read())
    fp.close()
    file_type = mimetypes.guess_type(file_name)
    if file_type is None:
        file_type = 'application/octet-stream'
    response['Content-Type'] = file_type
    response['Content-Length'] = str(os.stat(file_name).st_size)
    response['Content-Disposition'] = "attachment; filename=" + file_name_for_download
    return response


def create_ks14_docx(obj, act):
    base_app_dir = os.path.dirname(os.path.abspath(__file__))
    dynamic_dir_name = os.path.join(base_app_dir, 'dynamic')
    if not os.path.exists(dynamic_dir_name):
        os.mkdir(dynamic_dir_name)
    docx_template_dir = os.path.join(base_app_dir, 'docx_template')
    docx_template = os.path.join(docx_template_dir, 'KS14ActTemplate.docx')
    context = model_to_dict(obj, exclude=['id'])
    context_of_act = model_to_dict(act,exclude=['id'])
    context.update(context_of_act)
    doc = DocxTemplate(docx_template)
    doc.render(context)
    doc.save(os.path.join(dynamic_dir_name, "ks14.docx"))
    return get_file(os.path.join(dynamic_dir_name, "ks14.docx"), 'ks14.docx')


def create_ks3_xlsx (obj, act):
    base_app_dir = os.path.dirname(os.path.abspath(__file__))
    dynamic_dir_name = os.path.join(base_app_dir, 'dynamic')
    if not os.path.exists(dynamic_dir_name):
        os.mkdir(dynamic_dir_name)
    docx_template_dir = os.path.join(base_app_dir, 'docx_template')
    docx_template = os.path.join(docx_template_dir, 'KS3 Template.xlsx')
    wb = load_workbook(docx_template,read_only=False)
    ws = wb.active
    ws['B11'] = 'Общество с ограниченной ответственностью «' + obj.contractor + '»' + obj.contractor_address
    ws['J11'] = obj.contractor_OKPO
    ws['B13'] = 'Ремонт ' + act.system_genitive + ' в многоквартирном доме по адресу: ' + obj.address
    ws['J16'] = obj.contract_number
    s = obj.contract_date.split('.')
    ws['J17'] = s[0]
    ws['K17'] = s[1]
    ws['L17'] = s[2][0:4]
    ws['H24'] = act.act_number
    ws['I24'] = re.search(r'^[0-9][0-9].[0-9][0-9].[0-9][0-9][0-9][0-9]', act.act_date).group(0)
    ws['K24'] = re.search(r'^[0-9][0-9].[0-9][0-9].[0-9][0-9][0-9][0-9]', act.begin_fact_date).group(0)
    ws['L24'] = re.search(r'^[0-9][0-9].[0-9][0-9].[0-9][0-9][0-9][0-9]', act.act_date).group(0)
    ws['B29'] = 'Ремонт ' + act.system_genitive + ' в многоквартирном доме по адресу: ' + obj.address
    ws['K34'] = act.summa
    ws['A45'] = 'Инженер отдела строительного контроля №' + obj.supervisor_OSK_number
    ws['A46'] = 'по ' + obj.supervisor_decree_dative
    ws['K46'] = obj.supervisor_delegate
    ws['A49'] = 'Общество с ограниченной ответственностью «' + obj.contractor + '»'
    ws['K50'] = obj.contractor_delegate
    if obj.administration_order:
        ws['A53'] = obj.administration_order
    else:
        ws['A53'] = obj.owner_delegate_decree
        ws['K53'] = obj.owner_delegate
        ws['A54'] = '(должность)'
        ws['G54'] = '(подпись)'
        ws['K54'] = '  (расшифровка подписи)'
    ws['A56'] = 'Уполномоченный  представитель администрации '+ obj.district_genitive + ' района'
    ws['A57'] = obj.administration_delegate_position
    ws['A58'] = obj.administration_delegate_decree
    ws['K58'] = obj.administration_delegate
    wb.save(os.path.join(dynamic_dir_name, 'ks3.xlsx'))
    wb.close()
    return get_file(os.path.join(dynamic_dir_name, 'ks3.xlsx'), 'ks3.xlsx')


def create_ks2_xlsx (my_obj, my_act):
    base_app_dir = os.path.dirname(os.path.abspath(__file__))
    dynamic_dir_name = os.path.join(base_app_dir, 'dynamic')
    if not os.path.exists(dynamic_dir_name):
        os.mkdir(dynamic_dir_name)
    docx_template_dir = os.path.join(base_app_dir, 'docx_template')
    docx_template = os.path.join(docx_template_dir, 'KS2 Template.xlsx')
    wb = load_workbook(docx_template,read_only=False)
    ws = wb.active
    ws['C9'] = 'Общество с ограниченной ответственностью «' + my_obj.contractor + '»' + my_obj.contractor_address
    ws['Q9'] = my_obj.contractor_OKPO
    ws['C10'] = 'Ремонт ' + my_act.system_genitive + ' в многоквартирном доме по адресу: ' + my_obj.address
    ws['C11'] = 'Ремонт ' + my_act.system_genitive + ' в многоквартирном доме по адресу: ' + my_obj.address
    ws['Q13'] = my_obj.contract_number
    s = my_obj.contract_date.split('.')
    ws['Q14'] = s[0]
    ws['T14'] = s[1]
    ws['W14'] = s[2][0:4]
    ws['H21'] = my_act.act_number
    ws['K21'] = re.search(r'^[0-9][0-9].[0-9][0-9].[0-9][0-9][0-9][0-9]', my_act.act_date).group(0)
    ws['O21'] = re.search(r'^[0-9][0-9].[0-9][0-9].[0-9][0-9][0-9][0-9]', my_act.begin_fact_date).group(0)
    ws['R21'] = re.search(r'^[0-9][0-9].[0-9][0-9].[0-9][0-9][0-9][0-9]', my_act.act_date).group(0)
    ws['B23'] = 'Сметная (договорная) стоимость в соответствии с договором подряда (субподряда): ' + my_act.summa + 'руб.'
    ws['I30'] = my_act.summa
    ws['I32'] = my_act.summa
    ws['B34'] = 'Подрядчик: Общество с ограниченной ответственностью «' + my_obj.contractor + '»'
    ws['O35'] = my_obj.contractor_delegate
    ws['B46'] = 'Инженер отдела строительного контроля №' + my_obj.supervisor_OSK_number
    ws['B47'] = 'по ' + my_obj.supervisor_decree_dative
    ws['O47'] = my_obj.supervisor_delegate
    if my_obj.administration_order:
        ws['B50'] = my_obj.administration_order
    else:
        ws['B50'] = 'согласно ' + my_obj.owner_delegate_decree_genitive
        ws['O50'] = my_obj.owner_delegate
        ws['B51'] = '(должность)                                                       ' \
                    '                                                                   ' \
                    '                                                           (подпись)        (расшифровка подписи)'
    ws['B52'] = 'Уполномоченный  представитель администрации '+ my_obj.district_genitive + ' района'
    ws['B53'] = my_obj.administration_delegate_position
    ws['B54'] = my_obj.administration_delegate_decree
    ws['O54'] = my_obj.administration_delegate
    wb.save(os.path.join(dynamic_dir_name, 'ks2.xlsx'))
    wb.close()
    return get_file(os.path.join(dynamic_dir_name, 'ks2.xlsx'), 'ks2.xlsx')

def create_peresort(my_obj, my_act):
    base_app_dir = os.path.dirname(os.path.abspath(__file__))
    dynamic_dir_name = os.path.join(base_app_dir, 'dynamic')
    if not os.path.exists(dynamic_dir_name):
        os.mkdir(dynamic_dir_name)
    docx_template_dir = os.path.join(base_app_dir, 'docx_template')
    docx_template = os.path.join(docx_template_dir, 'PeresortTemplate.docx')
    context = model_to_dict(my_obj, exclude=['id'])
    context_act = model_to_dict(my_act, exclude=['id'])
    context.update(context_act)
    doc = DocxTemplate(docx_template)
    doc.render(context)
    doc.save(os.path.join(dynamic_dir_name, "act_peresort.docx"))
    return get_file(os.path.join(dynamic_dir_name, "act_peresort.docx"), 'act_peresort.docx')
