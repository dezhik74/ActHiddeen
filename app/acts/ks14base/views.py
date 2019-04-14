import mimetypes
import os
import re

from django.forms.models import model_to_dict
from django.shortcuts import render, get_object_or_404, redirect, HttpResponse
from django.views.generic import View
from docxtpl import DocxTemplate
from openpyxl import Workbook, load_workbook

from .models import *
from .forms import *

# Create your views here.

def ks14_list(request):
    objs = ObjectCommon.objects.order_by("-id")
    return render(request, 'ks14base/ks14list.html',
                  context={'objects': objs})


def copy_ks14(request, pk):
    my_obj = get_object_or_404(Ks14Act, pk=pk)
    s = my_obj.address
    if len(s) > 93:
        s = s[0:93]
    my_obj.address = 'Копия: ' + s
    my_obj.id = None
    my_obj.save()
    return redirect('ks14_list_url')


def delete_ks14(request, pk):
    my_obj = get_object_or_404(Ks14Act, pk=pk)
    my_obj.delete()
    return redirect('ks14_list_url')


def get_file(file_name, file_name_for_download):
    fp = open(file_name, "rb")
    response = HttpResponse(fp.read())
    fp.close()
    file_type = mimetypes.guess_type(file_name)
    if file_type is None:
        file_type = 'application/octet-stream'
    response['Content-Type'] = file_type
    response['Content-Length'] = str(os.stat(file_name).st_size)
    response['Content-Disposition'] = "attachment; filename=" + file_name_for_download
    return response


def make_ks14_docx (request, pk):
    base_app_dir = os.path.dirname(os.path.abspath(__file__))
    dynamic_dir_name = os.path.join(base_app_dir, 'dynamic')
    if not os.path.exists(dynamic_dir_name):
        os.mkdir(dynamic_dir_name)
    docx_template_dir = os.path.join(base_app_dir, 'docx_template')
    docx_template = os.path.join(docx_template_dir, 'KS14ActTemplate.docx')
    my_obj = get_object_or_404(Ks14Act, pk=pk)
    context = model_to_dict(my_obj, exclude=['id'])
    doc = DocxTemplate(docx_template)
    doc.render(context)
    doc.save(os.path.join(dynamic_dir_name, "ks14.docx"))
    return get_file(os.path.join(dynamic_dir_name, "ks14.docx"), 'ks14.docx')


def make_ks3_xlsx(request, pk):
    my_obj=get_object_or_404(Ks14Act, pk=pk)
    base_app_dir = os.path.dirname(os.path.abspath(__file__))
    dynamic_dir_name = os.path.join(base_app_dir, 'dynamic')
    if not os.path.exists(dynamic_dir_name):
        os.mkdir(dynamic_dir_name)
    docx_template_dir = os.path.join(base_app_dir, 'docx_template')
    docx_template = os.path.join(docx_template_dir, 'KS3 Template.xlsx')
    wb = load_workbook(docx_template,read_only=False)
    ws = wb.active
    ws['B11'] = 'Общество с ограниченной ответственностью «' + my_obj.contractor + '»' + my_obj.contractor_address
    ws['J11'] = my_obj.contractor_OKPO
    ws['B13'] = 'Капитальный ремонт ' + my_obj.system_genitive + ' в многоквартирном доме по адресу: ' + my_obj.address
    ws['J16'] = my_obj.contract_number
    s = my_obj.contract_date.split('.')
    ws['J17'] = s[0]
    ws['K17'] = s[1]
    ws['L17'] = s[2][0:4]
    ws['H24'] = my_obj.act_number
    ws['I24'] = re.search(r'^[0-9][0-9].[0-9][0-9].[0-9][0-9][0-9][0-9]', my_obj.act_date).group(0)
    ws['K24'] = re.search(r'^[0-9][0-9].[0-9][0-9].[0-9][0-9][0-9][0-9]', my_obj.begin_fact_date).group(0)
    ws['L24'] = re.search(r'^[0-9][0-9].[0-9][0-9].[0-9][0-9][0-9][0-9]', my_obj.act_date).group(0)
    ws['B29'] = 'Капитальный ремонт ' + my_obj.system_genitive + ' в многоквартирном доме по адресу: ' + my_obj.address
    ws['K34'] = my_obj.summa
    ws['A45'] = 'Инженер отдела строительного контроля №' + my_obj.supervisor_OSK_number
    ws['A46'] = 'по ' + my_obj.supervisor_decree_dative
    ws['K46'] = my_obj.supervisor_delegate
    ws['A49'] = 'Общество с ограниченной ответственностью «' + my_obj.contractor + '»'
    ws['K50'] = my_obj.contractor_delegate
    if my_obj.administration_order:
        ws['A53'] = my_obj.administration_order
    else:
        ws['A53'] = my_obj.owner_delegate_decree
        ws['K53'] = my_obj.owner_delegate
        ws['A54'] = '(должность)'
        ws['G54'] = '(подпись)'
        ws['K54'] = '  (расшифровка подписи)'
    ws['A56'] = 'Уполномоченный  представитель администрации '+ my_obj.district_prepositional + ' района'
    ws['A57'] = my_obj.administration_delegate_position
    ws['A58'] = my_obj.administration_delegate_decree
    ws['K58'] = my_obj.administration_delegate

    wb.save(os.path.join(dynamic_dir_name, 'ks3.xlsx'))
    wb.close()
    return get_file(os.path.join(dynamic_dir_name, 'ks3.xlsx'), 'ks3.xlsx')


def make_ks2_xlsx(request, pk):
    my_obj=get_object_or_404(Ks14Act, pk=pk)
    base_app_dir = os.path.dirname(os.path.abspath(__file__))
    dynamic_dir_name = os.path.join(base_app_dir, 'dynamic')
    if not os.path.exists(dynamic_dir_name):
        os.mkdir(dynamic_dir_name)
    docx_template_dir = os.path.join(base_app_dir, 'docx_template')
    docx_template = os.path.join(docx_template_dir, 'KS2 Template.xlsx')
    wb = load_workbook(docx_template,read_only=False)
    ws = wb.active
    ws['E9'] = 'Общество с ограниченной ответственностью «' + my_obj.contractor + '»' + my_obj.contractor_address
    ws['Q9'] = my_obj.contractor_OKPO
    ws['E10'] = 'Капитальный ремонт ' + my_obj.system_genitive + ' в многоквартирном доме по адресу: ' + my_obj.address
    ws['E11'] = 'Капитальный ремонт ' + my_obj.system_genitive + ' в многоквартирном доме по адресу: ' + my_obj.address
    ws['Q13'] = my_obj.contract_number
    s = my_obj.contract_date.split('.')
    ws['Q14'] = s[0]
    ws['T14'] = s[1]
    ws['W14'] = s[2][0:4]
    ws['H21'] = my_obj.act_number
    ws['K21'] = re.search(r'^[0-9][0-9].[0-9][0-9].[0-9][0-9][0-9][0-9]', my_obj.act_date).group(0)
    ws['O21'] = re.search(r'^[0-9][0-9].[0-9][0-9].[0-9][0-9][0-9][0-9]', my_obj.begin_fact_date).group(0)
    ws['R21'] = re.search(r'^[0-9][0-9].[0-9][0-9].[0-9][0-9][0-9][0-9]', my_obj.act_date).group(0)
    ws['B23'] = 'Сметная (договорная) стоимость в соответствии с договором подряда (субподряда): ' + my_obj.summa + 'руб.'
    ws['I30'] = my_obj.summa
    ws['I32'] = my_obj.summa
    ws['B34'] = 'Подрядчик: Общество с ограниченной ответственностью «' + my_obj.contractor + '»'
    ws['O35'] = my_obj.contractor_delegate
    ws['B46'] = 'Инженер отдела строительного контроля №' + my_obj.supervisor_OSK_number
    ws['B47'] = 'по ' + my_obj.supervisor_decree_dative
    ws['O47'] = my_obj.supervisor_delegate
    if my_obj.administration_order:
        ws['B50'] = my_obj.administration_order
    else:
        ws['B50'] = 'согласно ' + my_obj.owner_delegate_decree_genitive
        ws['O50'] = my_obj.owner_delegate
        ws['B51'] = '(должность)                                                       ' \
                    '                                                                   ' \
                    '                                                           (подпись)        (расшифровка подписи)'
    ws['B52'] = 'Уполномоченный  представитель администрации '+ my_obj.district_prepositional + ' района'
    ws['B53'] = my_obj.administration_delegate_position
    ws['B54'] = my_obj.administration_delegate_decree
    ws['O54'] = my_obj.administration_delegate

    wb.save(os.path.join(dynamic_dir_name, 'ks2.xlsx'))
    wb.close()
    return get_file(os.path.join(dynamic_dir_name, 'ks2.xlsx'), 'ks2.xlsx')


def make_peresort (request, pk):
    base_app_dir = os.path.dirname(os.path.abspath(__file__))
    dynamic_dir_name = os.path.join(base_app_dir, 'dynamic')
    if not os.path.exists(dynamic_dir_name):
        os.mkdir(dynamic_dir_name)
    docx_template_dir = os.path.join(base_app_dir, 'docx_template')
    docx_template = os.path.join(docx_template_dir, 'PeresortTemplate.docx')
    my_obj = get_object_or_404(Ks14Act, pk=pk)
    context = model_to_dict(my_obj, exclude=['id'])
    doc = DocxTemplate(docx_template)
    doc.render(context)
    doc.save(os.path.join(dynamic_dir_name, "act_peresort.docx"))
    return get_file(os.path.join(dynamic_dir_name, "act_peresort.docx"), 'act_peresort.docx')


class KS14_Detail(View):
    @staticmethod
    def get(request, pk):
        my_obj = get_object_or_404(ObjectCommon, pk=pk)
        return render(request, 'ks14base/ks14_detail.html',
                      context={'my_object': my_obj})


def ks14_object_edit(request, pk):
    my_obj = get_object_or_404(ObjectCommon, pk=pk)
    init_data = model_to_dict(my_obj, exclude=['id', 'acts'])
    if request.method == 'POST':
        my_form = ObjectCommonForm (request.POST)
        if my_form.is_valid():
            my_form.save()
            return redirect(my_obj)
        return render(request, 'ks14base/object_edit.html', context = {'my_obj' : my_obj,
                                                                        'my_form' : my_form})
    else:
        my_form = ObjectCommonForm(initial=init_data)
        return render(request, 'ks14base/object_edit.html', context={'my_obj': my_obj,
                                                                     'my_form': my_form})




